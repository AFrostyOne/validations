from openpyxl import load_workbook
import validations as val

wb = load_workbook('Inventory Workbook.xlsx')
ws = wb.active
ws_vendors = wb['Vendors']
desc_char_errors = ["Description has a 20 character limit. The following do not match:"]
SKU_form_errors = ["SKU must be in the format 12345-123. The following don't match:"]
price_errors = ["The price is invalid for the following items:"]
name_missing = ["Name is missing for the following items:"]
name_char = ["Item name has 15 character limit. The following are longer:"]
quantity_errors = ["The following quantities are invalid. They must be an integer."]
ID_missing = ["The following are missing the ID"]
ID_sequential = ["IDs must be sequential. The following IDs are not."]
vendor_compare_errors = ["The following vendors do not match a vendor from the vendor worksheet"]
invalid_email = ["The following are invalid email addresses"]
purchase_date_errors = ["The following dates are incorrect. Must be in mm/dd/yyyy format"]

master_error_lists = [desc_char_errors, SKU_form_errors, price_errors, name_missing, name_char, quantity_errors,
                      ID_missing, ID_sequential, vendor_compare_errors, invalid_email, purchase_date_errors]


# Check errors in column A - Item ID
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        ID_missing = val.list_missing(cell=cell, error_list=ID_missing)
        ID_sequential = val.list_sequential(cell=cell,
                                            error_list=ID_sequential)

# Check errors in column B - Item Name
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=2):
    for cell in row:
        name_missing = val.list_missing(cell=cell, error_list=name_missing,
                                        ID_col_offset=-1)
        name_char = val.list_max_character(cell=cell, max_char=15,
                                           error_list=name_char, ID_col_offset=-1)

# Check errors in column C - Item Price
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        price_errors = val.list_regex(cell=cell, reg="^\d*\.?\d{,2}$",
                                      error_list=price_errors, ID_col_offset=-2)

# Check errors in column D - Quantity
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=4, max_col=4):
    for cell in row:
        quantity_errors = val.list_regex(cell=cell, reg="^\d+$",
                                         error_list=quantity_errors, ID_col_offset=-3)

# Check errors in column E - SKU #
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=5, max_col=5):
    for cell in row:
        SKU_form_errors = val.list_regex(cell=cell, reg="^\d{5}-\d{3}$",
                                         error_list=SKU_form_errors, ID_col_offset=-4)

# Check errors in column F - Item Description
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=6, max_col=6):
    for cell in row:
        desc_char_errors = val.list_max_character(cell=cell, max_char=20,
                                                  error_list=desc_char_errors, ID_col_offset=-5)

# Check for errors in column G - Vendor
vendors = set()
for cell in ws_vendors['A']:
    vendors.add(cell.value)
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=7, max_col=7):
    for cell in row:
        vendor_compare_errors = val.from_list(cell=cell,
                                              error_list=vendor_compare_errors, compare_set=vendors,
                                              ID_col_offset=-6)

# Check for errors in column H - Vendor Email
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=8, max_col=8):
    for cell in row:
        invalid_email = val.list_regex(cell=cell,
                                       reg="^([a-zA-Z0-9_\-\.]+)@([a-zA-Z0-9_\-\.]+)\.([a-zA-Z]{2,5})$",
                                       error_list=invalid_email, ID_col_offset=-7)

# Check for errors in column I - Last Purchase Date
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=9, max_col=9):
    for cell in row:
        purchase_date_errors = val.list_regex(cell=cell,
                                              reg="^(0[1-9]|1[012])[ /.](0[1-9]|[12][0-9]|3[01])[ /.](19|20)\d\d$",
                                              error_list=purchase_date_errors, ID_col_offset=-8)

print(val.create_error_instructions(purchase_date_errors))

print("ID Missing", ID_missing)
print("ID Sequential", ID_sequential)
print("SKU errors", SKU_form_errors)
print("prices errors", price_errors)
print("quantity errors", quantity_errors)
print("name missing", name_missing)
print("vendor_compare_errors", vendor_compare_errors)
print("Invalid email", invalid_email)
print("Purchase Date Errors,", purchase_date_errors)


for error_list in master_error_lists:
    val.color_cells(worksheet=ws, dict_list=error_list[1:len(error_list)])
wb.save("Inventory Workbook-Validated.xlsx")


with open("Clarifications.txt", 'a') as f:
    for error_list in master_error_lists:
        message = val.create_error_instructions(error_list)
        if message:
            message = "\n\n" + message
            f.write(message)

