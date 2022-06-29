# This module contains functions that can be used to validate excel workbook
import re
from openpyxl.styles import PatternFill


def dict_cell(cell, ID_row_offset=0, ID_col_offset=0):
    cell_dictionary = {"row": cell.row, "column": cell.column,
                       "value": str(cell.value), "ID": str(cell.offset(ID_row_offset, ID_col_offset).value)}
    return cell_dictionary


def list_max_character(cell, max_char, error_list, ID_row_offset=0,
                       ID_col_offset=0):
    """Search cell for max character limit. Create dictionary if
    the cell doesn't pass and adds to list."""
    if len(str(cell.value)) > max_char:
        error_list.append(dict_cell(cell, ID_row_offset, ID_col_offset))
    return error_list


def list_regex(cell, reg, error_list, ID_row_offset=0,
               ID_col_offset=0):
    reg_check = re.compile(reg)
    if not reg_check.match(str(cell.value)):
        error_list.append(dict_cell(cell, ID_row_offset, ID_col_offset))
    return error_list


def list_missing(cell, error_list, ID_row_offset=0, ID_col_offset=0):
    if not cell.value:
        error_list.append(dict_cell(cell, ID_row_offset, ID_col_offset))
    return error_list


def list_sequential(cell, error_list, ID_row_offset=0, ID_col_offset=0):
    if not cell.offset(-1, 0).value:
        error_list.append(dict_cell(cell, ID_row_offset, ID_col_offset))
        return error_list
    if not cell.value == cell.offset(-1, 0).value + 1:
        error_list.append(dict_cell(cell, ID_row_offset, ID_col_offset))
    return error_list


def from_list(cell, error_list, compare_set, ID_row_offset=0, ID_col_offset=0):
    if not cell.value in compare_set:
        error_list.append(dict_cell(cell, ID_row_offset, ID_col_offset))
    return error_list


def create_error_instructions(error_list):
    if len(error_list) < 2:
        return
    error_message = error_list[0]
    for error in error_list[1:len(error_list)]:
        error_message = \
            f"{error_message} \n value: {error['value']} ID: {error['ID']} In row {error['row']} column " \
            + f"{error['column']}"
        # error_message = error_message + "\n value: " + str(error['value']) + "ID: " \
        # + str(error['ID']) + "In row " + str(error['row']), "In column" + str(error['column'])
    return error_message


def color_cells(worksheet, dict_list, color="FF0000"):
    for cell_dict in dict_list:
        cell = worksheet.cell(row=int(cell_dict['row']), column=int(cell_dict['column']))
        cell.fill = PatternFill("solid", start_color=color)


def validate_max_character(worksheet, max_char, start_row, start_col, \
                           end_row=10000, end_col=500, ID_col_offset=0):
    """Search specified cells for max character limit. Creates list of dictionaries
    with the cells that don't pass and highlights them red."""
    char_errors = []
    if end_row > worksheet.max_row:
        end_row = worksheet.max_row
    if end_col > worksheet.max_column:
        end_col = worksheet.max_column
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row,
                                   min_col=start_col, max_col=end_col):
        for cell in row:
            if not cell.value:
                continue
            if len(str(cell.value)) > max_char:
                char_errors.append({"row": cell.row, "column": cell.column,
                                    "value": cell.value, "ID": cell.offset(0, ID_col_offset).value})
    return char_errors
